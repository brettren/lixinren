# !/usr/bin/python
# -*- coding: UTF-8 -*-
import csv
import datetime
import json
import pandas as pd
import requests
import config
import openpyxl
import xlsxwriter
from openpyxl.styles import PatternFill, colors
from openpyxl.utils import column_index_from_string
from openpyxl.styles import numbers

token = config.token
metricsList = [
    "pe_ttm.y10.mcw",
    "pb.y10.mcw",
    "dyr.mcw"
]
stockMapsCN = {

}
stockMapsHK = {

}
all_index_drawdown_csv_file = "all_index_drawdown_test.csv"
excel_file = "all_index_drawdown_test.xlsx"


class Stock:
    def __int__(self, code, pe, pePos, peToLowest, pb, pbPos, pbToLowest, roe, roe_vs_pe, div):
        self.code = code
        self.pe = pe
        self.pePos = pePos
        self.peToLowest = peToLowest
        self.pb = pb
        self.pbPos = pbPos
        self.pbToLowest = pbToLowest
        self.roe = roe
        self.roe_vs_pe = roe_vs_pe
        self.div = div


# endDate = datetime.datetime.today().strftime('%Y-%m-%d')
endDate = config.today
startDate = (datetime.datetime.today() - datetime.timedelta(weeks=52 * 10)).strftime('%Y-%m-%d')
headers = {'Content-Type': 'application/json'}
summary = list()
summary.append(
    [endDate, "", "", "", "style", "position", "Current drawdown", "Largest drawdown", "To lowest drawdown 📉🔽",
     "To Highest required 📈🔼", "pe", "pe10pos", "pe to lowest", "pb", "pb10pos", "pb to lowest", "roe", "roe/pe", "div",
     "funds"])


def get_index_codes(country):
    stock_codes_url = f'https://open.lixinger.com/api/{country}/index'
    stock_codes_data = {
        "token": token,
    }
    response = requests.post(url=stock_codes_url, headers=headers, data=json.dumps(stock_codes_data)).json()
    data = response["data"]
    cnName_map = {}
    publishDate_map = {}
    for d in data:
        cnName_map[d["stockCode"]] = d["name"]
        publishDate_map[d["stockCode"]] = d["launchDate"][0:10]
    return cnName_map, publishDate_map


def get_response(url, data):
    response = requests.post(url=url, headers=headers, data=json.dumps(data)).json()
    return response


def getFundamental(stockCodes, country):
    valueMap = {}
    i = 0
    while i < len(stockCodes):
        getFundamentalSubPart(valueMap, stockCodes[i: i + 100], country)
        i += 100
    return valueMap


def getFundamentalSubPart(valueMap, stockCodes, country):
    response = get_response(url=f'https://open.lixinger.com/api/{country}/index/fundamental',
                            data={
                                "stockCodes": stockCodes,
                                "metricsList": metricsList,
                                "token": token,
                                "date": endDate
                            })
    for stock in response["data"]:
        index = Stock()
        index.code = stock["stockCode"]
        pe = stock["pe_ttm.y10.mcw.cv"]
        index.pe = "{:.2}".format(pe)
        index.pePos = stock["pb.y10.mcw.cvpos"]
        index.peToLowest = "{:.2%}".format(
            1 - stock["pe_ttm.y10.mcw.minv"] / stock["pe_ttm.y10.mcw.cv"])
        index.pb = "{:.2}".format(stock["pb.y10.mcw.cv"])
        index.pbPos = stock["pb.y10.mcw.cvpos"]
        index.pbToLowest = "{:.2%}".format(1 - stock["pb.y10.mcw.minv"] / stock["pb.y10.mcw.cv"])
        roe = float(index.pb) / float(index.pe)
        index.roe = "{:.2%}".format(roe)
        index.roe_vs_pe = "{:.2}".format(roe / pe * 100) if pe > 0 else ""
        if "dyr.mcw" in stock:
            dividend = stock["dyr.mcw"]
            index.div = "{:.2%}".format(dividend)
        valueMap[stock["stockCode"]] = index
    return valueMap


def getEachIndexDrawdown(sector, funds, fund_styles, code, publishDate, valueInfo, value, country):
    # filter the index without tracking fund
    if funds.get(code) is None:
        print(f"{code} needs update tracking fund")
        return
    if not funds[code].strip():
        print(f"{code} has no tracking fund, no need to get drawdown data")
        return
    try:
        response = get_response(url=f"https://open.lixinger.com/api/{country}/index/drawdown",
                                data={
                                    "stockCode": code,
                                    "startDate": startDate,
                                    "endDate": endDate,
                                    "token": token,
                                    "granularity": "y3",
                                })
        if 'message' in response and response['message'] == 'success':
            data = response['data']
            current_drawdown = data[0]['value']
            currentDrawdownText = "{:.2%}".format(current_drawdown)
            # sort the data by date
            sorted_data = sorted(data, key=lambda day: day['value'])
            # get the lowest and highest index
            largest_drawdown = sorted_data[0]['value']
            drawdown = "{:.2%}".format(largest_drawdown)
            number = 1 - (1 + largest_drawdown) / (1 + current_drawdown)
            toLowest = "{:.2%}".format(number)
            toHighest = "{:.2%}".format(1 / (1 + current_drawdown) - 1)
            hasPosition = "*" if code in config.positionMaps.keys() else ""
            try:
                summary.append(
                    [sector, code, publishDate, value, fund_styles.get(code), hasPosition, currentDrawdownText,
                     drawdown,
                     toLowest, toHighest, valueInfo.pe, "{:.2%}".format(valueInfo.pePos), valueInfo.peToLowest,
                     valueInfo.pb, "{:.2%}".format(valueInfo.pbPos), valueInfo.pbToLowest, valueInfo.roe,
                     valueInfo.roe_vs_pe, valueInfo.div, funds[code]])
            except Exception as e:
                print(str(e))
        else:
            print("Failed to fetch data")
    except:
        print(f"Failed to get {code} drawdown data")



def mark_as_percent(df, column):
    # Convert the percentage strings to percentage numbers.
    df[column] = df[column].str.replace("%", "")
    df[column] = df[column].astype(float)
    df[column] = df[column].div(100)


def convert_to_excel():
    # Load the CSV file into a pandas DataFrame
    df = pd.read_csv(all_index_drawdown_csv_file)

    mark_as_percent(df, "Current drawdown")
    mark_as_percent(df, "Largest drawdown")
    mark_as_percent(df, "To lowest drawdown 📉🔽")
    mark_as_percent(df, "To Highest required 📈🔼")
    mark_as_percent(df, "pe10pos")
    mark_as_percent(df, "pe to lowest")
    mark_as_percent(df, "pb10pos")
    mark_as_percent(df, "pb to lowest")
    mark_as_percent(df, "roe")
    mark_as_percent(df, "div")

    writer = pd.ExcelWriter(excel_file, engine='xlsxwriter')
    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name="Sheet1", index=False)

    # get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # add format (percentage with 2 decimals)
    percent_fmt = workbook.add_format({'num_format': '0.00%'})
    # set column to percentage format
    worksheet.set_column(6, 9, None, percent_fmt)
    worksheet.set_column(11, 12, None, percent_fmt)
    worksheet.set_column(14, 16, None, percent_fmt)
    worksheet.set_column(18, 18, None, percent_fmt)
    writer.close()

    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file)

    # Select the worksheet
    worksheet = workbook['Sheet1']

    # Set the fill color for column B to green
    green_fill = PatternFill(start_color='81e3a8', end_color='81e3a8', fill_type='solid')
    for cell in worksheet['I:I']:
        cell.fill = green_fill
    blue_fill = PatternFill(start_color='77edeb', end_color='77edeb', fill_type='solid')
    for cell in worksheet['M:M']:
        cell.fill = blue_fill
    orange_fill = PatternFill(start_color='fa7e4d', end_color='fa7e4d', fill_type='solid')
    for cell in worksheet['Q:Q']:
        cell.fill = orange_fill

    # Add a filter for rows 1 and 2
    worksheet.auto_filter.ref = 'A1:T1'

    # Freeze the top row
    worksheet.freeze_panes = worksheet['A1']

    # Save the changes to the Excel file
    workbook.save(excel_file)


def fetch_all_index_drawdown():
    # Load the CSV file into a pandas DataFrame
    funds = {}
    with open("all_index_tracking_fund.csv", 'r', encoding='utf_8_sig') as csv_file:
        csv_reader = csv.reader(csv_file)
        for row in csv_reader:
            funds[row[0]] = row[1]
    fund_styles = {}
    with open("all_index_style.csv", 'r', encoding='utf_8_sig') as csv_file:
        csv_reader = csv.reader(csv_file)
        for row in csv_reader:
            if not funds.get(row[0]):
                print(f"{row[0]} has no tracking fund, no need to add style")
                continue
            fund_styles[row[0]] = row[2]
    stockMapsCN, publishDate_map = get_index_codes("cn")
    valueMap = getFundamental(list(stockMapsCN.keys()), "cn")
    for key, value in stockMapsCN.items():
        if valueMap.get(key) is None:
            print(f"{key} cannot get Fundamental info")
            continue
        getEachIndexDrawdown("A股", funds, fund_styles, key, publishDate_map[key], valueMap[key], value, "cn")
    stockMapsHK, publishDate_map = get_index_codes("hk")
    valueMap = getFundamental(list(stockMapsHK.keys()), "hk")
    for key, value in stockMapsHK.items():
        getEachIndexDrawdown("港股", funds, fund_styles, key, publishDate_map[key], valueMap[key], value, "hk")
    summary[1:] = sorted(summary[1:], key=lambda index: float(index[8].replace("%", "")))
    with open(all_index_drawdown_csv_file, 'w', encoding='utf_8_sig') as summaryf:
        writer = csv.writer(summaryf)
        writer.writerows(summary)
    # for line in summary:
    #     print(line)


fetch_all_index_drawdown()
convert_to_excel()
