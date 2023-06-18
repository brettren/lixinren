#!/usr/bin/python
# -*- coding:utf8 -*-

import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import datetime

url = 'https://open.lixinger.com/api/a/indice/fundamental'

headers = {'Content-Type': 'application/json'}

today = datetime.date.today().strftime("%Y-%m-%d")
# today = "2018-10-12"

data = {
    "stockCodes": [
        "000016",
        "000300",
        "399904",
        "000905",
        "000952",
        "399958",
        "1000002",
        "000991",
        "000922",
        "000827",
        "000852",
        "000932",
        "000992",
        "000993",
        "399975",
        "399005",
        "399006",
        "399673",
        "399812",
        "399967",
        "399986",
        "399971"
    ],
    "metrics": [
        "pe_ttm.y_10.weightedAvg",
        "pb.y_10.weightedAvg"
    ],
    "token": "afb329e1-73ea-41d6-9672-dcf5ba22a4bc",
    "date": today
}

response = requests.post(url=url, headers=headers, data=json.dumps(data)).json()
print 'response:'
print response


def process(data, sheet):
    row = 2
    info_col = 1
    date_col = 1
    sheet.cell(row=row, column=date_col).value = data[0]["date"]
    info_col = info_col + 1
    for d in data:
        # sheet.cell(row=row, column=info_col).value = d["stockCnName"]
        sub_process(d, "pe_ttm", "weightedAvg", row, info_col)
        info_col = info_col + 1
        sub_process(d, "pb", "weightedAvg", row, info_col)
        info_col = info_col + 1


def sub_process(d, pe_pb_type, avg, row, col):
    pe_ttm = d[pe_pb_type]
    y_10 = pe_ttm["y_10"]
    val = y_10[avg]["latestVal"]
    sheet.cell(row=row, column=col).value = val
    risk_val = y_10[avg]["riskVal"]
    chance_val = y_10[avg]["chanceVal"]
    if val >= risk_val:
        sheet.cell(row=row, column=col).font = ft_red
    elif val <= chance_val:
        sheet.cell(row=row, column=col).font = ft_green


def input_code(data, row, col):
    for d in data:
        sheet.cell(row=row, column=col).value = d["stockCode"]
        col = col + 2


workbook_ = load_workbook(u"stock_stat.xlsx")

sheet_names = workbook_.sheetnames

sheet = workbook_[sheet_names[3]]

data = response["data"]

ft_red = Font(color=colors.RED)
ft_green = Font(color=colors.GREEN)

# input_code(data, 1, 3)
process(data, sheet)

workbook_.save(u"stock_stat.xlsx")
