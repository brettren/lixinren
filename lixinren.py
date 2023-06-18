#!/usr/bin/python
# -*- coding:utf8 -*-

import sys
import requests
import json
from openpyxl import load_workbook
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
import datetime


ft_red = Font(color=colors.COLOR_INDEX[2])
ft_darkRed = Font(color=colors.COLOR_INDEX[16])
ft_darkGreen = Font(color=colors.COLOR_INDEX[17])
ft_green = Font(color=colors.COLOR_INDEX[3])
metrics = [
        "pe_ttm.y10.mcw",
        "pe_ttm.y5.mcw",
        "pe_ttm.y10.ew",
        "pe_ttm.y5.ew",
        "pb.y10.mcw",
        "pb.y5.mcw",
        "pb.y10.ew",
        "pb.y5.ew",
        "dyr.mcw",
        "dyr.ew",
        "cp"  # for index
]

stockCodes = [
    "000016",
    "000300",
    "000904",
    "000905",
    "000952",
    "000958",
    "1000002",
    "000991",
    "000922",
    "000827",
    "000852",
    "000932",
    "000992",
    "000993",
    "399975",
    "399005",
    "399006",
    "399673",
    "399812",
    "399967",
    "399986",
    "399971",
    "000933",
    "399989",
    "000807",
    "399806",
    "000935",
    "000989",
    "399997",
    "000931",
    "000934",
    "399998",
    "399970",
    "000986",
    "000929",
    "000987",
    "000928",
    "000988",
    "000930",
    "000937",
    "000995",
    "000994",
    "000936",
    "000808",
    "000990",
    "399976",
    "000015",
    "000978",
    "399974",
    "399809",
    "000979",
    "399394",
    "399324",
    "000919",
    "930782",
    "950090",
    "H11136",
]
token = "632bbc8f-bf24-4bd0-b7e9-29cdd3b20200"
headers = {'Content-Type': 'application/json'}
# today = datetime.date.today().strftime("%Y-%m-%d")
today = "2023-02-17"
if len(sys.argv) > 1:
    today = sys.argv[1]

col_name_define_row = 1
begin_row = 2

info_col = 1
type_col = info_col + 1
current_col = type_col + 1
lowest_y10_col = current_col + 1
falling_to_lowest_y10_col = lowest_y10_col + 1
highest_y10_col = falling_to_lowest_y10_col + 1
rising_to_highest_y10_col = highest_y10_col + 1
rising_to_chance_y10_col = rising_to_highest_y10_col + 1
rising_to_median_y10_col = rising_to_chance_y10_col + 1
rising_to_risk_y10_col = rising_to_median_y10_col + 1
highest_to_lowest_y10_col = rising_to_risk_y10_col + 1
lowest_y5_col = highest_to_lowest_y10_col + 1
falling_to_lowest_y5_col = lowest_y5_col + 1
highest_y5_col = falling_to_lowest_y5_col + 1
rising_to_highest_y5_col = highest_y5_col + 1
highest_to_lowest_y5_col = rising_to_highest_y5_col + 1
roe_weighted_col = highest_to_lowest_y5_col + 1
roe_equal_col = roe_weighted_col + 1
dividend_col = roe_equal_col + 1
close_point_col = dividend_col + 1


def get_response(url, data):
    response = requests.post(url=url, headers=headers, data=json.dumps(data)).json()
    return response


def process(data, sheet, isGZ):
    if len(data) == 0:
        return
    sheet.cell(row=col_name_define_row, column=info_col).value = data[0]["date"]

    sheet.cell(row=col_name_define_row, column=current_col).value = "current"
    sheet.cell(row=col_name_define_row, column=lowest_y10_col).value = "lowest_y10"
    sheet.cell(row=col_name_define_row, column=falling_to_lowest_y10_col).value = "falling_to_lowest_y10"
    sheet.cell(row=col_name_define_row, column=highest_y10_col).value = "highest_y10"
    sheet.cell(row=col_name_define_row, column=rising_to_highest_y10_col).value = "rising_to_highest_y10"
    sheet.cell(row=col_name_define_row, column=rising_to_chance_y10_col).value = "rising_to_chance_y10_col"
    sheet.cell(row=col_name_define_row, column=rising_to_median_y10_col).value = "rising_to_median_y10_col"
    sheet.cell(row=col_name_define_row, column=rising_to_risk_y10_col).value = "rising_to_risk_y10_col"
    sheet.cell(row=col_name_define_row, column=highest_to_lowest_y10_col).value = "highest_to_lowest_y10"
    sheet.cell(row=col_name_define_row, column=lowest_y5_col).value = "lowest_y5"
    sheet.cell(row=col_name_define_row, column=falling_to_lowest_y5_col).value = "falling_to_lowest_y5"
    sheet.cell(row=col_name_define_row, column=highest_y5_col).value = "highest_y5"
    sheet.cell(row=col_name_define_row, column=rising_to_highest_y5_col).value = "rising_to_highest_y5"
    sheet.cell(row=col_name_define_row, column=highest_to_lowest_y5_col).value = "highest_to_lowest_y5"
    sheet.cell(row=col_name_define_row, column=roe_weighted_col).value = "roe_weighted"
    sheet.cell(row=col_name_define_row, column=roe_equal_col).value = "roe_equal"
    sheet.cell(row=col_name_define_row, column=dividend_col).value = "dividend(weighted/equal)"
    sheet.cell(row=col_name_define_row, column=close_point_col).value = "close_point"

    row = begin_row
    for d in data:
        # sheet.cell(row=row, column=info_col).value = d["stockCnName"]
        # if the code order modified in api, uncomment below
        # sheet.cell(row=row, column=info_col).value = cnName_map[d["stockCode"]]
        sheet.cell(row=row, column=type_col).value = "pe weighted"
        sub_process(d, "pe_ttm", "mcw", row)
        if "mcw" in d["dyr"]:
            dividend = d["dyr"]["mcw"]
        else:
            dividend = None
        sheet.cell(row=row, column=dividend_col).value = dividend

        if "cp" in d:
            close_point = d["cp"]
            sheet.cell(row=row, column=close_point_col).value = close_point
        row = row + 1

        # next row
        sheet.cell(row=row, column=info_col).value = d["stockCode"]
        sheet.cell(row=row, column=type_col).value = "pe equal"
        sub_process(d, "pe_ttm", "ew", row)

        if "ew" in d["dyr"]:
            dividend = d["dyr"]["ew"]
        else:
            dividend = None
        sheet.cell(row=row, column=dividend_col).value = dividend

        row = row + 1

        # next row
        # sheet.cell(row=row, column=info_col).value = d["publishDate"]
        # if the code order modified in api, uncomment below
        # sheet.cell(row=row, column=info_col).value = publishDate_map[d["stockCode"]]
        sheet.cell(row=row, column=type_col).value = "pb weighted"
        sub_process(d, "pb", "mcw", row)
        row = row + 1

        # next row
        sheet.cell(row=row, column=type_col).value = "pb equal"
        sub_process(d, "pb", "ew", row)
        row = row + 2

    row = begin_row
    for d in data:
        if sheet.cell(row=row, column=current_col).value != None:
            sheet.cell(row=row, column=roe_weighted_col).value = sheet.cell(row=row+2, column=current_col).value/sheet.cell(row=row, column=current_col).value
        else:
            sheet.cell(row=row, column=roe_weighted_col).value = None
        if sheet.cell(row=row+1, column=current_col).value != None:
            sheet.cell(row=row, column=roe_equal_col).value = sheet.cell(row=row+3, column=current_col).value/sheet.cell(row=row+1, column=current_col).value
        else:
            sheet.cell(row=row, column=roe_equal_col).value = None
        row = row + 5


def sub_process(d, pe_pb_type, avg, row):
    pe_ttm = d[pe_pb_type]
    y10 = pe_ttm["y10"]
    val = y10[avg]["cv"]
    sheet.cell(row=row, column=current_col).value = val
    set_color(y10[avg], val, row, current_col)

    val = y10[avg]["minv"]
    sheet.cell(row=row, column=lowest_y10_col).value = val
    if sheet.cell(row=row, column=current_col).value != None:
        sheet.cell(row=row, column=falling_to_lowest_y10_col).value \
            = 1 - sheet.cell(row=row, column=lowest_y10_col).value / sheet.cell(row=row, column=current_col).value
    else:
        sheet.cell(row=row, column=falling_to_lowest_y10_col).value = None
    val = y10[avg]["maxv"]
    sheet.cell(row=row, column=highest_y10_col).value = val
    if sheet.cell(row=row, column=current_col).value != None:
        sheet.cell(row=row, column=rising_to_highest_y10_col).value \
            = sheet.cell(row=row, column=highest_y10_col).value / sheet.cell(row=row, column=current_col).value - 1
    else:
        sheet.cell(row=row, column=rising_to_highest_y10_col).value = None
    val = y10[avg]["q2v"]
    if sheet.cell(row=row, column=current_col).value != None:
        sheet.cell(row=row, column=rising_to_chance_y10_col).value \
            = val / sheet.cell(row=row, column=current_col).value - 1
    else:
        sheet.cell(row=row, column=rising_to_chance_y10_col).value = None
    val = y10[avg]["q5v"]
    if sheet.cell(row=row, column=current_col).value != None:
        sheet.cell(row=row, column=rising_to_median_y10_col).value \
            = val / sheet.cell(row=row, column=current_col).value - 1
    else:
        sheet.cell(row=row, column=rising_to_median_y10_col).value = None
    val = y10[avg]["q8v"]
    if sheet.cell(row=row, column=current_col).value != None:
        sheet.cell(row=row, column=rising_to_risk_y10_col).value \
            = val / sheet.cell(row=row, column=current_col).value - 1
    else:
        sheet.cell(row=row, column=rising_to_risk_y10_col).value = None
    sheet.cell(row=row, column=highest_to_lowest_y10_col).value \
        = sheet.cell(row=row, column=highest_y10_col).value / sheet.cell(row=row, column=lowest_y10_col).value

    y5 = pe_ttm["y5"]
    val = y5[avg]["minv"]
    sheet.cell(row=row, column=lowest_y5_col).value = val
    if sheet.cell(row=row, column=current_col).value != None:
        sheet.cell(row=row, column=falling_to_lowest_y5_col).value \
            = 1 - sheet.cell(row=row, column=lowest_y5_col).value / sheet.cell(
            row=row, column=current_col).value
    else:
        sheet.cell(row=row, column=falling_to_lowest_y5_col).value = None
    val = y5[avg]["maxv"]
    sheet.cell(row=row, column=highest_y5_col).value = val
    if sheet.cell(row=row, column=current_col).value != None:
        sheet.cell(row=row, column=rising_to_highest_y5_col).value \
            = sheet.cell(row=row, column=highest_y5_col).value / sheet.cell(
            row=row, column=current_col).value - 1
    else:
        sheet.cell(row=row, column=rising_to_highest_y5_col).value = None
    sheet.cell(row=row, column=highest_to_lowest_y5_col).value \
        = sheet.cell(row=row, column=highest_y5_col).value / sheet.cell(
        row=row, column=lowest_y5_col).value


def set_color(avg, val, row, col):
    risk_val = avg["q8v"]
    chance_val = avg["q2v"]
    median_val = avg["q5v"]
    if val >= risk_val:
        sheet.cell(row=row, column=col).font = ft_red
    elif median_val <= val < risk_val:
        sheet.cell(row=row, column=col).font = ft_darkRed
    elif chance_val <= val < median_val:
        sheet.cell(row=row, column=col).font = ft_darkGreen
    elif val < chance_val:
        sheet.cell(row=row, column=col).font = ft_green

# with open("stock_stat.json", 'r') as load_f:
#     response = json.load(load_f)


def get_stock_codes(level):
    stock_codes_url = 'https://open.lixinger.com/api/a/industry'
    stock_codes_data = {
        "token": token,
        "source": "cni",
        "level": level
    }
    response = requests.post(url=stock_codes_url, headers=headers, data=json.dumps(stock_codes_data)).json()
    data = response["data"]
    code_list = []
    cnName_map = {}
    for d in data:
        code_list.append(d["stockCode"])
        cnName_map[d["stockCode"]] = d["name"]
    return code_list, cnName_map


def get_index_codes():
    stock_codes_url = 'https://open.lixinger.com/api/a/index'
    stock_codes_data = {
        "token": token,
        "stockCodes": stockCodes
    }
    response = requests.post(url=stock_codes_url, headers=headers, data=json.dumps(stock_codes_data)).json()
    data = response["data"]
    cnName_map = {}
    publishDate_map = {}
    for d in data:
        cnName_map[d["stockCode"]] = d["name"]
        publishDate_map[d["stockCode"]] = d["launchDate"]
    return cnName_map, publishDate_map


workbook_ = load_workbook(u"stock_stat.xlsx")
sheet_names = workbook_.sheetnames

sheet = workbook_[sheet_names[0]]
cnName_map, publishDate_map = get_index_codes()
response = get_response(url='https://open.lixinger.com/api/a/index/fundamental',
                        data={
                            "stockCodes": stockCodes,
                            "metricsList": metrics,
                            "token": token,
                            "date": today
                        })
# print 'response a:'
# print response
data = response["data"]
process(data, sheet, False)
publishDate_map.clear()

sheet = workbook_[sheet_names[1]]
response = get_response(url='https://open.lixinger.com/api/h/index/fundamental',
                        data={
                            "stockCodes": [
                                "HSI",
                                "HSCEI"
                            ],
                            "metricsList": metrics,
                            "token": token,
                            "date": today
                        })
# print 'response hk:'
# print response
data = response["data"]
process(data, sheet, False)

sheet = workbook_[sheet_names[2]]
code_list, cnName_map = get_stock_codes("one")
response = get_response(url='https://open.lixinger.com/api/a/industry/fundamental/cni',
                        data={
                            "stockCodes": code_list,
                            "metricsList": metrics,
                            "token": token,
                            "date": today
                        })
# print 'response gz one:'
# print response
data = response["data"]
process(data, sheet, True)

sheet = workbook_[sheet_names[3]]
code_list, cnName_map = get_stock_codes("two")
response = get_response(url='https://open.lixinger.com/api/a/industry/fundamental/cni',
                        data={
                            "stockCodes": code_list,
                            "metricsList": metrics,
                            "token": token,
                            "date": today
                        })
# print 'response gz two:'
# print response
data = response["data"]
process(data, sheet, True)

sheet = workbook_[sheet_names[4]]
code_list, cnName_map = get_stock_codes("three")
response = get_response(url='https://open.lixinger.com/api/a/industry/fundamental/cni',
                        data={
                            "stockCodes": code_list,
                            "metricsList": metrics,
                            "token": token,
                            "date": today
                        })
# print 'response gz three:'
# print response
data = response["data"]
process(data, sheet, True)

workbook_.save(u"stock_stat.xlsx")
